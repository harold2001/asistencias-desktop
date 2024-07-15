"""Main controller"""

import sqlite3
import subprocess
from copy import deepcopy
from datetime import datetime, timedelta
from tkinter import Frame, Entry, Tk, filedialog, Menu
from ttkbootstrap.toast import ToastNotification
from ttkbootstrap.constants import END, PRIMARY, INFO, YES, BOTH, SUCCESS, DANGER
from ttkbootstrap.dialogs.dialogs import Messagebox
from ttkbootstrap.tableview import Tableview
from ttkbootstrap.scrolled import ScrolledFrame
from openpyxl import Workbook
import ttkbootstrap as ttk


class Main:
    """Main program"""

    def __init__(self, window_app, db_app):
        self.wind = window_app
        self.db_name = db_app
        self.input_codigo = None
        self.main_frame = None

        # Ventana minizada: Ancho y alto de la pantalla a la mitad
        width = self.wind.winfo_screenwidth() / 2
        height = self.wind.winfo_screenheight() / 2
        self.wind.geometry(f"{int(width)}x{int(height)}")

        # Ventanta maximizada: Full
        self.wind.state("zoomed")

        self.wind.title("Sistema de asistencias")
        menubar = Menu(self.wind)

        # Crear menús si aún no existen
        if not menubar.index("end"):
            # Asistencias
            asistencias_menu = Menu(menubar, tearoff=0)
            asistencias_menu.add_command(
                label="Marcar entrada", command=self.set_principal_view
            )
            asistencias_menu.add_command(
                label="Marcar salida", command=self.set_salida_view
            )
            menubar.add_cascade(label="Asistencias", menu=asistencias_menu)

            # Reportes
            reportes_menu = Menu(menubar, tearoff=0)
            reportes_menu.add_command(
                label="General", command=self.set_reporte_general_view
            )
            reportes_menu.add_command(
                label="Por alumno", command=self.set_reporte_alumno_view
            )
            reportes_menu.add_command(
                label="Por grado y sección", command=self.set_reporte_grado_view
            )
            menubar.add_cascade(label="Reportes", menu=reportes_menu)

            self.wind.config(menu=menubar)

            # Alumnos
            reportes_menu = Menu(menubar, tearoff=0)
            reportes_menu.add_command(
                label="Crear nuevo", command=self.set_alumno_add_view
            )
            reportes_menu.add_command(label="Ver todos", command=self.set_alumnos_view)
            menubar.add_cascade(label="Alumnos", menu=reportes_menu)

            self.wind.config(menu=menubar)

        self.set_principal_view()

    def reset_view(self, main_title, is_expand=True, padding=20):
        """Eliminar todos los widgets de la vista actual"""
        # Eliminar widgets de la vista principal
        for widget in self.wind.winfo_children():
            if widget.winfo_class() != "Menu":
                widget.destroy()

        # Quitar el escuchador de eventos "Enter" de la ventana principal
        self.wind.unbind("<Return>")

        # Frame Container
        self.main_frame = ttk.Frame(
            self.wind, borderwidth=0, relief="flat", padding=padding
        )
        self.main_frame.pack(expand=is_expand)

        # Titulo
        main_titulo = ttk.Label(
            self.main_frame,
            text=main_title,
            font=("Sans-serif", 26),
        )
        main_titulo.pack(expand=True)

    def reset_input_codigo(self, event):
        """Ctrl + BackSpace -> Borrar todo el contenido del input"""
        print(event)
        self.input_codigo.delete(0, END)

    def run_query(self, query, parameters=()):
        """Ejecutar cualquier query y obtener el resultado"""
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            result = cursor.execute(query, parameters)
        return result

    def display_success_toast(self, message):
        """Notificación para cuando una operación se realice con éxito"""
        toast = ToastNotification(
            title="Operación exitosa", message=message, duration=2000, icon=None
        )
        toast.show_toast()

    def display_error_box(self, message, parent=None):
        """Mostrar un Messagebox centrado"""
        if parent is None:
            parent = self.main_frame

        Messagebox.show_error(message=message, title="Advertencia", parent=parent)
        # overlay.destroy()

    def set_change_view_link_corner(self, content, on_click):
        """Colocar link button en la esquina superior derecha"""
        # Botón "Ver reporte" posicionado en la esquina superior derecha
        boton_reporte = ttk.Button(
            self.wind, text=content, bootstyle=INFO, command=on_click
        )
        boton_reporte.place(relx=0.0, y=25, x=30, anchor="nw")

    def register_asistencia(self, codigo_alumno):
        """Función para registrar asistencia cuando se presione 'Enter' en la vista principal"""
        self.input_codigo.delete(0, END)
        if codigo_alumno in (None, ""):
            return self.display_error_box("Código inválido")

        alumno = self.run_query(
            "SELECT alumno_id, nombres, apellido_paterno FROM alumnos WHERE codigo = ?",
            [codigo_alumno],
        ).fetchone()

        if alumno in (None, "") or len(alumno) == 0:
            return self.display_error_box("No se encontró el alumno")

        now = datetime.now()
        fecha = now.strftime("%Y-%m-%d")

        existentes = self.run_query(
            """
            SELECT
                an.asistencia_id 
            FROM
                asistencias an
            INNER JOIN alumnos al ON
                an.alumno_id = al.alumno_id
            WHERE al.codigo = ? AND an.fecha=?
            """,
            [codigo_alumno, fecha],
        ).fetchall()

        if len(existentes) > 0:
            return self.display_error_box(
                "Hoy ya se marcó la entrada de este alumno. Puedes marcar su salida."
            )

        hora_entrada = now.strftime("%H:%M:%S")
        asistencia = self.run_query(
            "INSERT INTO asistencias(alumno_id, hora_entrada, fecha) VALUES (?, ?, ?)",
            [alumno[0], hora_entrada, fecha],
        )

        if asistencia:
            return self.display_success_toast(
                f"Asistencia marcada para el alumno: {alumno[1]} {alumno[2]}"
            )

        return self.display_error_box("Error interno al registrar la entrada")

    def register_salida(self, codigo_alumno):
        """Marcar salida de un alumno"""
        self.input_codigo.delete(0, END)

        if codigo_alumno in (None, ""):
            return self.display_error_box("Código inválido")

        alumno = self.run_query(
            "SELECT alumno_id, nombres, apellido_materno FROM alumnos WHERE codigo = ?",
            [codigo_alumno],
        ).fetchone()

        if alumno in (None, "") or len(alumno) == 0:
            return self.display_error_box("No se encontró el alumno")

        now = datetime.now()
        fecha = now.strftime("%Y-%m-%d")

        existentes = self.run_query(
            """
            SELECT
                an.hora_salida 
            FROM
                asistencias an
            INNER JOIN alumnos al ON
                an.alumno_id = al.alumno_id
            WHERE al.codigo = ? AND an.fecha=?
            """,
            [codigo_alumno, fecha],
        ).fetchall()

        if existentes in (None, "") or len(existentes) == 0:
            return self.display_error_box(
                "Primero debes marcar entrada para este alumno"
            )

        if existentes[0][0] not in (None, ""):
            return self.display_error_box("Ya se marcó la salida de este alumno")

        hora_salida = now.strftime("%H:%M:%S")
        salida = self.run_query(
            "UPDATE asistencias SET hora_salida=? WHERE alumno_id = ? AND fecha = ?",
            [hora_salida, alumno[0], fecha],
        )

        if salida:
            return self.display_success_toast(
                f"Salida marcada para el alumno: {alumno[1]} {alumno[2]}"
            )

        return self.display_error_box("Error interno al registrar la salida")

    def create_alumno(
        self, codigo, nombres, paterno, materno, fecha, grado_id, seccion
    ):
        """Crear alumno"""
        detalle_grado = self.run_query(
            "SELECT detalle_grado_id FROM detalle_grados WHERE grado_id=? AND seccion=?",
            [grado_id, seccion],
        ).fetchone()

        fecha_formatted = datetime.strptime(fecha, "%d-%m-%Y").strftime("%Y-%m-%d")
        alumno = self.run_query(
            "INSERT INTO alumnos(codigo, nombres, apellido_paterno, apellido_materno, fecha_ingreso, detalle_grado_id) VALUES (?, ?, ?, ?, ?, ?)",
            [codigo, nombres, paterno, materno, fecha_formatted, detalle_grado[0]],
        )

        if alumno:
            self.set_alumno_add_view()
            return self.display_success_toast("Alumno creado con éxito")

        return self.display_error_box("Error interno al crear un alumno")

    def update_alumno(
        self, alumno_id, codigo, nombres, paterno, materno, fecha, grado_id, seccion
    ):
        """Editar alumno"""
        detalle_grado = self.run_query(
            "SELECT detalle_grado_id FROM detalle_grados WHERE grado_id=? AND seccion=?",
            [grado_id, seccion],
        ).fetchone()

        fecha_formatted = datetime.strptime(fecha, "%d-%m-%Y").strftime("%Y-%m-%d")
        alumno = self.run_query(
            "UPDATE alumnos SET codigo=?, nombres=?, apellido_paterno=?, apellido_materno=?, fecha_ingreso=?, detalle_grado_id=? WHERE alumno_id=?",
            [
                codigo,
                nombres,
                paterno,
                materno,
                fecha_formatted,
                detalle_grado[0],
                alumno_id,
            ],
        )

        if alumno:
            self.set_alumnos_view()
            return self.display_success_toast("Alumno actualizado con éxito")

        return self.display_error_box("Error interno al actualizar el alumno")

    def export_to_excel(
        self, table, default_name, alumno_grado=None, alumno_seccion=None
    ):
        """Exportar datos de una tabla TableView en formato Excel"""
        # Traer datos de las cabeceras y registros de la tabla
        headers = [col.headertext for col in table.tablecolumns]
        records = [row.values for row in table.tablerows]

        if len(headers) == 0 or len(records) == 0:
            self.display_error_box("No hay datos para exportar en esta tabla")
            return

        # Obtener la ubicación y el nombre del archivo del usuario
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            initialfile=f"{default_name}.xlsx",
        )
        if not file_path:
            self.display_error_box("Ruta inválida para guardar un archivo")
            return

        wb = Workbook()
        ws = wb.active

        # Celda A1
        ws.cell(row=1, column=1, value="ALUMNO:")
        # Combinar las celdas B1, C1 y D1
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)
        ws.cell(row=1, column=2, value=default_name)

        # Celda B1 y B2
        ws.cell(row=2, column=1, value="GRADO:")
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=4)
        ws.cell(row=2, column=2, value=alumno_grado)

        # Celda C1 y C2
        ws.cell(row=3, column=1, value="SECCIÓN:")
        ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=4)
        ws.cell(row=3, column=2, value=alumno_seccion)

        # Escribir los encabezados de las columnas
        for col, column_name in enumerate(headers, start=1):
            ws.cell(row=4, column=col, value=column_name)

        # Escribir los datos de las filas
        for row, row_data in enumerate(records, start=5):
            for col, data in enumerate(row_data, start=1):
                ws.cell(row=row, column=col, value=data)

        # Guardar el archivo Excel en la ubicación especificada
        try:
            wb.save(file_path)
        except PermissionError:
            self.display_error_box(
                "No se pudo guardar el archivo. Permiso denegado. Cierre el archivo."
            )
            return

        # Abrir el archivo después de guardarlo
        answer = Messagebox.show_question(
            message="Archivo guardado con éxito. ¿Desea abrirlo?",
            title="Operación exitosa",
            alert=True,
            parent=self.main_frame,
            buttons=["No:secondary", "Sí:primary"],
        )

        if answer.lower() == "sí":
            subprocess.Popen([file_path], shell=True)
        else:
            Messagebox.show_info(
                message=f"Archivo guardado en la ruta {file_path}",
                title="Archivo guardado",
                alert=True,
                parent=self.main_frame,
            )

    def export_to_excel_2(self, table, fecha, alumno_grado=None, alumno_seccion=None):
        """Exportar datos de una tabla TableView en formato Excel"""
        # Traer datos de las cabeceras y registros de la tabla
        headers = [col.headertext for col in table.tablecolumns]
        records = [row.values for row in table.tablerows]

        if len(headers) == 0 or len(records) == 0:
            self.display_error_box("No hay datos para exportar en esta tabla")
            return

        # Obtener la ubicación y el nombre del archivo del usuario
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            initialfile=f"{fecha}.xlsx",
        )
        if not file_path:
            self.display_error_box("Ruta inválida para guardar un archivo")
            return

        wb = Workbook()
        ws = wb.active

        # Celda A1
        ws.cell(row=1, column=1, value="FECHA:")
        # Combinar las celdas B1, C1 y D1
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)
        ws.cell(row=1, column=2, value=fecha)

        # Celda B1 y B2
        ws.cell(row=2, column=1, value="GRADO:")
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=4)
        ws.cell(row=2, column=2, value=alumno_grado)

        # Celda C1 y C2
        ws.cell(row=3, column=1, value="SECCIÓN:")
        ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=4)
        ws.cell(row=3, column=2, value=alumno_seccion)

        # Escribir los encabezados de las columnas
        for col, column_name in enumerate(headers, start=1):
            ws.cell(row=4, column=col, value=column_name)

        # Escribir los datos de las filas
        for row, row_data in enumerate(records, start=5):
            for col, data in enumerate(row_data, start=1):
                ws.cell(row=row, column=col, value=data)

        # Guardar el archivo Excel en la ubicación especificada
        try:
            wb.save(file_path)
        except PermissionError:
            self.display_error_box(
                "No se pudo guardar el archivo. Permiso denegado. Cierre el archivo."
            )
            return

        # Abrir el archivo después de guardarlo
        answer = Messagebox.show_question(
            message="Archivo guardado con éxito. ¿Desea abrirlo?",
            title="Operación exitosa",
            alert=True,
            parent=self.main_frame,
            buttons=["No:secondary", "Sí:primary"],
        )

        if answer.lower() == "sí":
            subprocess.Popen([file_path], shell=True)
        else:
            Messagebox.show_info(
                message=f"Archivo guardado en la ruta {file_path}",
                title="Archivo guardado",
                alert=True,
                parent=self.main_frame,
            )

    def export_to_excel_3(self, tables, grado=None, seccion=None):
        """Exportar datos de una tabla TablheView en formato Excel"""
        # Obtener la ubicación y el nombre del archivo del usuario
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            initialfile=f"{grado}.xlsx",
        )
        if not file_path:
            self.display_error_box("Ruta inválida para guardar un archivo")
            return

        wb = Workbook()
        ws = wb.active

        # Celda A1
        ws.cell(row=1, column=1, value="SECCIÓN:")
        # Combinar las celdas B1, C1 y D1
        ws.cell(row=1, column=2, value=seccion)

        # Celda B1 y B2
        ws.cell(row=2, column=1, value="GRADO:")
        ws.cell(row=2, column=2, value=grado)

        # Escribir los datos de las filas
        first_row = 3
        for mes, table in tables:
            # Traer datos de las cabeceras y registros de la tabla
            records = [row.values for row in table.tablerows]
            headers = [col.headertext for col in table.tablecolumns]
            first_row += 1
            ws.cell(row=first_row, column=1, value="MES:")
            ws.cell(row=first_row, column=2, value=mes)
            first_row += 1

            # Escribir los encabezados de las columnas
            for col, column_name in enumerate(headers, start=1):
                ws.cell(row=first_row, column=col, value=column_name)
            first_row += 1

            for row, row_data in enumerate(records, start=first_row):
                for col, data in enumerate(row_data, start=1):
                    ws.cell(row=row, column=col, value=data)
                first_row += 1

        # Guardar el archivo Excel en la ubicación especificada
        try:
            wb.save(file_path)
        except PermissionError:
            self.display_error_box(
                "No se pudo guardar el archivo. Permiso denegado. Cierre el archivo."
            )
            return

        # Abrir el archivo después de guardarlo
        answer = Messagebox.show_question(
            message="Archivo guardado con éxito. ¿Desea abrirlo?",
            title="Operación exitosa",
            alert=True,
            parent=self.main_frame,
            buttons=["No:secondary", "Sí:primary"],
        )

        if answer.lower() == "sí":
            subprocess.Popen([file_path], shell=True)
        else:
            Messagebox.show_info(
                message=f"Archivo guardado en la ruta {file_path}",
                title="Archivo guardado",
                alert=True,
                parent=self.main_frame,
            )

    def set_principal_view(self):
        """Mostrar vista principal"""
        self.reset_view("Esperando código de barras...")
        # Frame para el input y el botón
        frame_input_boton = Frame(self.main_frame)
        frame_input_boton.pack(expand=True)

        # Input código de barras
        self.input_codigo = Entry(
            frame_input_boton, font=("Helvetica", 18), justify="center"
        )
        self.input_codigo.pack(side="left")
        self.input_codigo.focus()

        # Evento Ctrl + Delete
        self.input_codigo.bind("<Control-BackSpace>", self.reset_input_codigo)

        # Botón "agregar manualmente"
        boton_agregar = ttk.Button(
            frame_input_boton,
            text="Agregar manualmente",
            bootstyle=PRIMARY,
            command=lambda: self.register_asistencia(self.input_codigo.get().upper()),
        )
        boton_agregar.pack(side="left", padx=10, pady=20)

        # Escuchador de evento "Enter" a la ventana principal
        self.wind.bind(
            "<Return>",
            lambda event: self.register_asistencia(self.input_codigo.get().upper()),
        )

    def set_salida_view(self):
        """Vista para marcar salida"""
        self.reset_view("Marcar salida")
        # Frame para el input y el botón
        frame_input_boton = Frame(self.main_frame)
        frame_input_boton.pack(expand=True)

        # Input código de barras
        self.input_codigo = Entry(
            frame_input_boton, font=("Helvetica", 18), justify="center"
        )
        self.input_codigo.pack(side="left")
        self.input_codigo.focus()

        # Evento Ctrl + Delete
        self.input_codigo.bind("<Control-BackSpace>", self.reset_input_codigo)

        # Botón "agregar manualmente"
        boton_agregar = ttk.Button(
            frame_input_boton,
            text="Agregar manualmente",
            bootstyle=PRIMARY,
            command=lambda: self.register_salida(self.input_codigo.get().upper()),
        )
        boton_agregar.pack(side="left", padx=10, pady=20)

        # Escuchador de evento "Enter" a la ventana principal
        self.wind.bind(
            "<Return>",
            lambda event: self.register_salida(self.input_codigo.get().upper()),
        )

    def set_reporte_general_view(self):
        """Mostrar segunda vista con todos los grados disponibles"""
        self.reset_view("Reporte general mensual")

        # Debajo del título
        ttk.Label(
            self.main_frame,
            text="Selecciona un grado y sección:",
            font=("Sans-Serif", 11),
        ).pack(fill="x", pady=(30, 20))

        # Grados
        grados = self.run_query(
            "SELECT grado_id, grado FROM grados WHERE grado_id >= 7"
        ).fetchall()
        grados_dict = {grado: grado_id for grado_id, grado in grados}
        combobox_grados = ttk.Combobox(
            self.main_frame,
            bootstyle="primary",
            values=list(grados_dict.keys()),
            state="readonly",
        )
        combobox_grados.set("Grado")
        combobox_grados.pack(pady=20, padx=20, side="left")

        # Secciones
        secciones = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
        combobox_secciones = ttk.Combobox(
            self.main_frame, bootstyle="primary", values=secciones, state="readonly"
        )
        combobox_secciones.set("Sección")
        combobox_secciones.pack(pady=20, padx=20, side="left")

        def set_validate_report():
            """Validar los datos y mostrar el reporte"""
            grado = combobox_grados.get()
            seccion = combobox_secciones.get()
            if grado == "Grado" or seccion == "Sección":
                self.display_error_box("Selecciona un grado y sección")
                return

            self.set_reporte_general_table(
                grados_dict[grado],
                grado,
                seccion,
            )

        ttk.Button(
            self.main_frame,
            text="Buscar",
            bootstyle=PRIMARY,
            padding=(20, 10),
            command=set_validate_report,
        ).pack(expand=True, fill="x")

    def set_reporte_general_table(self, grado_id, grado, seccion):
        """Mostrar tabla con datos de la búsqueda por grado y seccion de manera mensual"""
        self.reset_view(grado, is_expand=False)
        self.set_change_view_link_corner(
            "Volver al buscador", self.set_reporte_general_view
        )

        # Datos del grado
        sf_tablas = ScrolledFrame(self.wind, autohide=True)
        sf_tablas.pack(fill=BOTH, expand=YES, padx=0, pady=0)

        ttk.Label(
            self.main_frame,
            text=f"Sección: {seccion}",
            font=("Sans-serif", 11),
            justify="center",
            anchor="nw",
        ).pack(expand=True, fill="x")

        meses = [
            "Enero",
            "Febrero",
            "Marzo",
            "Abril",
            "Mayo",
            "Junio",
            "Julio",
            "Agosto",
            "Septiembre",
            "Octubre",
            "Noviembre",
            "Diciembre",
        ]
        year = datetime.now().year
        dias_es = ["L", "M", "M", "J", "V"]

        # Datos de los alumnos
        alumnos = self.run_query(
            """
            SELECT
                al.alumno_id,
                al.nombres,
                al.apellido_paterno,
                al.apellido_materno
            FROM
                alumnos al
            INNER JOIN detalle_grados dg ON
                al.detalle_grado_id = dg.detalle_grado_id
            WHERE
                dg.grado_id = ?
                AND dg.seccion = ?;
            """,
            [grado_id, seccion],
        ).fetchall()

        # Tabla por mes
        tables = []
        for mes_num, mes_nombre in enumerate(meses, start=1):
            ttk.Label(
                sf_tablas,
                text=f"{mes_nombre} {year}",
                font=("Sans-serif", 14),
                justify="center",
                anchor="center",
            ).pack(expand=True, fill="x")

            coldata = [
                {"text": "N°", "stretch": True},
                {"text": "Nombres y apellidos", "stretch": True},
            ]

            primer_dia = datetime(year, mes_num, 1)
            ultimo_dia = (
                (datetime(year, mes_num + 1, 1) - timedelta(days=1))
                if mes_num != 12
                else datetime(year + 1, 1, 1) - timedelta(days=1)
            )

            # Dias del mes en iteracion
            dias = {}
            current_day = primer_dia
            while current_day <= ultimo_dia:
                if current_day.weekday() < 5:
                    dia_fecha = current_day.day
                    dia_letra = dias_es[current_day.weekday()]
                    dias[dia_fecha] = {"letra": dia_letra, "asistencia": "I"}
                    coldata.append(
                        {"text": f"{dia_letra}-{dia_fecha}", "stretch": True}
                    )
                current_day += timedelta(days=1)
            dias_cantidad = len(dias.keys())
            coldata.append({"text": "Asistencias", "stretch": True})
            coldata.append({"text": "% Asistencia", "stretch": True})
            coldata.append({"text": "Inasistencias", "stretch": True})
            alumnos_rows = list(alumnos)
            mes_num_string = f"{mes_num:02d}"

            rowsdata = []

            # Generar datos de asistencia de cada alumno por mes
            for numero, alumno in enumerate(alumnos_rows, start=1):
                alumno_mes_asistencias = self.run_query(
                    """
                    SELECT
                        an.hora_entrada,
                        an.hora_salida,
                        an.fecha,
                        al.alumno_id,
                        al.nombres,
                        al.apellido_paterno,
                        al.apellido_materno
                    FROM
                        alumnos al
                    LEFT JOIN asistencias an ON
                        an.alumno_id = al.alumno_id
                    INNER JOIN detalle_grados dg ON
                        al.detalle_grado_id = dg.detalle_grado_id
                    WHERE
                        dg.grado_id = ?
                        AND dg.seccion = ?
                        AND strftime('%m', an.fecha) = ?
                        AND al.alumno_id = ?;
                    """,
                    [grado_id, seccion, mes_num_string, alumno[0]],
                ).fetchall()

                alumno_asistencias_cantidad = len(alumno_mes_asistencias)
                alumno_dias = deepcopy(dias)

                for asistencia in alumno_mes_asistencias:
                    asistencia_fecha = datetime.strptime(asistencia[2], "%Y-%m-%d")
                    dia = asistencia_fecha.day
                    if dia in alumno_dias:
                        alumno_dias[dia]["asistencia"] = "A"

                letras = [valor["asistencia"] for valor in alumno_dias.values()]
                data = [numero, f"{alumno[1]} {alumno[2]} {alumno[3]}"]
                data.extend(letras)
                data.append(alumno_asistencias_cantidad)
                asistencia_porcentaje = (
                    alumno_asistencias_cantidad / dias_cantidad * 100
                )
                data.append(f"{asistencia_porcentaje:.1f}%")
                data.append(dias_cantidad - alumno_asistencias_cantidad)
                rowsdata.append(tuple(data))

            # Crear tabla
            dt = Tableview(
                master=sf_tablas,
                coldata=coldata,
                rowdata=rowsdata,
                paginated=True,
                pagesize=40,
                searchable=False,
                bootstyle=PRIMARY,
                autofit=True,
                autoalign=True,
                height=40,
            )
            dt.pack(fill=BOTH, expand=True, padx=40, pady=(0, 50))
            tables.append((mes_nombre, dt))

            # Centrar cabeceras y filas de la tabla
            for col_id in dt.view["columns"]:
                if int(col_id) >= 2:
                    dt.view.column(col_id, anchor="center")
                    dt.view.heading(col_id, anchor="center")

        export_button = ttk.Button(
            self.wind,
            text="Exportar a Excel",
            bootstyle=SUCCESS,
            command=lambda: self.export_to_excel_3(tables, grado, seccion),
        )
        export_button.place(relx=1.0, y=25, x=-30, anchor="ne")

    def set_reporte_alumno_view(self):
        """Buscar un alumno por nombre y mostrar su reporte de asistencias en otra vista"""
        self.reset_view("Buscar por alumno")

        # Debajo del título
        ttk.Label(
            self.main_frame,
            text="Escribe el nombre de un alumno y luego dale click para ver su reporte:",
            font=("Sans-Serif", 11),
        ).pack(fill="x", pady=(30, 20))

        # Frame para el Entry y el botón
        frame_input_boton = Frame(self.main_frame)
        frame_input_boton.pack(expand=True)

        # Frame para el input y resultados
        frame_entry_list = Frame(frame_input_boton)
        frame_entry_list.pack(side="left", fill="x", padx=(0, 10))

        # Input nombre del alumno
        nombre_alumno = Entry(
            frame_entry_list,
            font=("Helvetica", 18),
            justify="center",
        )
        nombre_alumno.config(width=60)
        nombre_alumno.focus()
        nombre_alumno.pack(expand=True)

        # Lista de opciones para mostrar los resultados
        frame_resultados = ScrolledFrame(frame_entry_list, padding=0, height=300)
        frame_resultados.pack(fill="x", expand=YES)
        frame_resultados.hide_scrollbars()
        frame_resultados.disable_scrolling()

        # Evento KeyPress
        nombre_alumno.bind(
            "<KeyRelease>",
            lambda event: self.set_reporte_alumno_opciones(
                nombre_alumno.get(), frame_resultados
            ),
        )

        # Evento Ctrl + Delete
        nombre_alumno.bind(
            "<Control-BackSpace>", lambda event: nombre_alumno.delete(0, END)
        )

    def set_reporte_alumno_opciones(self, nombre, resultados_frame):
        """Mostrar opciones debajo del buscador"""
        # Limpiar resultados anteriores
        for widget in resultados_frame.winfo_children():
            widget.destroy()

        if nombre:
            match = self.run_query(
                """
                SELECT
                    a.alumno_id,
                    a.codigo,
                    a.nombres,
                    a.apellido_paterno,
                    a.apellido_materno,
                    g.grado,
                    dg.seccion
                FROM
                    alumnos a
                INNER JOIN detalle_grados dg 
                    ON
                    dg.detalle_grado_id = a.detalle_grado_id
                INNER JOIN grados g 
                    ON
                    g.grado_id = dg.grado_id
                WHERE
                    nombres LIKE ?
                    OR apellido_paterno LIKE ?
                    OR apellido_materno LIKE ?
            """,
                [f"%{nombre}%", f"%{nombre}%", f"%{nombre}%"],
            ).fetchall()

            stylebtn = ttk.Style()
            stylebtn.configure("Custom.TButton", font=("Sans-Serif", 11))

            for alumno in match:
                ttk.Button(
                    resultados_frame,
                    text=f'{alumno[2]} {alumno[3]} {alumno[4]} - {alumno[5]} "{alumno[6]}"',
                    bootstyle=INFO,
                    command=lambda alumno=alumno: self.set_reporte_alumno_table(alumno),
                    style="Custom.TButton",
                    padding=10,
                ).pack(fill="x", expand=True)
            resultados_frame.show_scrollbars()
            resultados_frame.enable_scrolling()
        else:
            resultados_frame.config(height=0)
            resultados_frame.hide_scrollbars()
            resultados_frame.disable_scrolling()

    def set_reporte_alumno_table(self, alumno):
        """Mostrar vista con el reporte de asistencia del alumno especificado"""
        self.reset_view("Reporte por alumno")
        self.set_change_view_link_corner(
            "Volver al buscador", self.set_reporte_alumno_view
        )
        alumno_nombre = f"{alumno[2]} {alumno[3]} {alumno[4]}"

        # Datos del alumno
        datos_frame = ttk.Frame(self.main_frame, width=80)
        datos_frame.pack(expand=True, fill="x", padx=0, pady=(30, 20))

        ttk.Label(
            datos_frame,
            text=f"Alumno: {alumno_nombre}",
            font=("Sans-serif", 11),
            justify="left",
            anchor="w",
        ).pack(expand=True, fill="x")

        ttk.Label(
            datos_frame,
            text=f"Grado: {alumno[5]}",
            font=("Sans-serif", 11),
            justify="left",
            anchor="w",
        ).pack(expand=True, fill="x")

        ttk.Label(
            datos_frame,
            text=f"Sección: {alumno[6]}",
            font=("Sans-serif", 11),
            justify="left",
            anchor="w",
        ).pack(expand=True, fill="x")

        # Asistencias del alumno seleccionado
        asistencias = self.run_query(
            "SELECT hora_entrada, hora_salida, fecha FROM asistencias WHERE alumno_id = ?",
            [alumno[0]],
        ).fetchall()

        if len(asistencias) == 0:
            ttk.Label(
                self.main_frame,
                text="No existen asistencias registradas para este alumno",
                font=("Sans-serif", 15),
                bootstyle="danger",
                padding=(0, 30),
                justify="center",
            ).pack(expand=True, fill="x")
            return

        coldata = [
            {"text": "Año", "stretch": True},
            {"text": "Mes", "stretch": True},
            {"text": "Día", "stretch": True},
            {"text": "Entrada", "stretch": True},
            {"text": "Salida", "stretch": True},
        ]

        months_in_spanish = {
            1: "Enero",
            2: "Febrero",
            3: "Marzo",
            4: "Abril",
            5: "Mayo",
            6: "Junio",
            7: "Julio",
            8: "Agosto",
            9: "Septiembre",
            10: "Octubre",
            11: "Noviembre",
            12: "Diciembre",
        }
        rowdata = []

        for asistencia in asistencias:
            entrada = datetime.strptime(asistencia[0], "%H:%M:%S")

            if asistencia[1] not in (None, ""):
                salida = datetime.strptime(asistencia[1], "%H:%M:%S").strftime(
                    "%I:%M:%S %p"
                )
            else:
                salida = ""

            fecha = datetime.strptime(asistencia[2], "%Y-%m-%d")
            row = (
                fecha.year,
                months_in_spanish[fecha.month],
                fecha.day,
                entrada.strftime("%I:%M:%S %p"),
                salida,
            )
            rowdata.append(row)

        ttk.Label(
            self.main_frame,
            text="Escribe y presiona ENTER para buscar",
            font=("Sans-serif", 10),
            justify="left",
            anchor="nw",
        ).pack(expand=True, fill="x")

        # Crear tabla
        dt = Tableview(
            master=self.main_frame,
            coldata=coldata,
            rowdata=rowdata,
            paginated=True,
            searchable=True,
            bootstyle=PRIMARY,
            autofit=True,
            autoalign=True,
        )
        dt.pack(fill=BOTH, expand=YES, padx=10, pady=10)

        # Centrar cabeceras y filas de la tabla
        for col_id in dt.view["columns"]:
            dt.view.column(col_id, anchor="center")
            dt.view.heading(col_id, anchor="center")

        # Crear el botón de exportar
        export_button = ttk.Button(
            self.main_frame,
            text="Exportar a Excel",
            command=lambda: self.export_to_excel(
                dt, alumno_nombre, alumno[5], alumno[6]
            ),
            bootstyle=SUCCESS,
        )
        export_button.pack(pady=10)

    def set_reporte_grado_view(self):
        """Vista del reporte por grado, seccion y fecha"""
        self.reset_view("Reporte por grado y sección")

        # Debajo del título
        ttk.Label(
            self.main_frame,
            text="Selecciona un grado, sección y fecha para ver un reporte específico:",
            font=("Sans-Serif", 11),
        ).pack(fill="x", pady=(30, 20))

        # Grados
        grados = self.run_query(
            "SELECT grado_id, grado FROM grados WHERE grado_id >= 7"
        ).fetchall()
        grados_dict = {grado: grado_id for grado_id, grado in grados}
        combobox_grados = ttk.Combobox(
            self.main_frame,
            bootstyle="primary",
            values=list(grados_dict.keys()),
            state="readonly",
        )
        combobox_grados.set("Grado")
        combobox_grados.pack(pady=20, padx=20, side="left")

        # Secciones
        secciones = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
        combobox_secciones = ttk.Combobox(
            self.main_frame, bootstyle="primary", values=secciones, state="readonly"
        )
        combobox_secciones.set("Sección")
        combobox_secciones.pack(pady=20, padx=20, side="left")

        # Fecha
        date_entry = ttk.DateEntry(
            self.main_frame, bootstyle="primary", dateformat="%d-%m-%Y"
        )
        date_entry.pack(expand=True, side="left", padx=20, pady=20)

        def set_validate_report():
            """Validar los datos y mostrar el reporte"""
            grado = combobox_grados.get()
            seccion = combobox_secciones.get()
            fecha = date_entry.entry.get()
            if grado == "Grado" or seccion == "Sección":
                self.display_error_box("Selecciona un grado y sección")
                return

            self.set_reporte_grado_table(
                grados_dict[grado],
                grado,
                seccion,
                fecha,
            )

        ttk.Button(
            self.main_frame,
            text="Buscar",
            bootstyle=PRIMARY,
            padding=(20, 10),
            command=set_validate_report,
        ).pack(expand=True, fill="x")

    def set_reporte_grado_table(self, grado_id, grado, seccion, fecha):
        """Mostrar tabla con datos de la búsqueda por grado, seccion y fecha"""
        self.reset_view(grado)
        self.set_change_view_link_corner(
            "Volver al buscador", self.set_reporte_grado_view
        )

        fecha_datetime = datetime.strptime(fecha, "%d-%m-%Y")

        # Datos del grado
        datos_frame = ttk.Frame(self.main_frame, width=100)
        datos_frame.pack(expand=True, fill="x", padx=0, pady=(30, 20))

        ttk.Label(
            datos_frame,
            text=f"Sección: {seccion}",
            font=("Sans-serif", 11),
            justify="left",
            anchor="nw",
        ).pack(expand=True, fill="x")

        ttk.Label(
            datos_frame,
            text=f"Fecha: {fecha}",
            font=("Sans-serif", 11),
            justify="left",
            anchor="nw",
        ).pack(expand=True, fill="x")

        # Asistencias del grado y seccion seleccionado
        asistencias = self.run_query(
            """
            SELECT
                an.hora_entrada,
                an.hora_salida,
                al.codigo,
                al.nombres,
                al.apellido_paterno,
                al.apellido_materno
            FROM
                asistencias an
            INNER JOIN alumnos al ON
                an.alumno_id = al.alumno_id
            INNER JOIN detalle_grados dg ON
                al.detalle_grado_id = dg.detalle_grado_id
            WHERE
                dg.grado_id = ?
                AND dg.seccion = ?
                AND an.fecha = ?;
            """,
            [grado_id, seccion, fecha_datetime.strftime("%Y-%m-%d")],
        ).fetchall()

        if len(asistencias) == 0:
            ttk.Label(
                self.main_frame,
                text="No existen asistencias registradas para esta fecha, sección y grado juntos.",
                font=("Sans-serif", 15),
                bootstyle="danger",
                padding=(0, 30),
                justify="center",
            ).pack(expand=True, fill="x")
            return

        coldata = [
            {"text": "Código", "stretch": True},
            {"text": "Nombre", "stretch": True},
            {"text": "Hora de entrada", "stretch": True},
            {"text": "Hora de salida", "stretch": True},
            {"text": "Fecha", "stretch": True},
        ]
        rowdata = []

        for asistencia in asistencias:
            entrada = datetime.strptime(asistencia[0], "%H:%M:%S")

            if asistencia[1] is None:
                salida = ""
            else:
                salida = datetime.strptime(asistencia[1], "%H:%M:%S").strftime(
                    "%I:%M %p"
                )

            row = (
                asistencia[2],
                f"{asistencia[3]} {asistencia[4]} {asistencia[5]}",
                entrada.strftime("%I:%M %p"),
                salida,
                fecha,
            )
            rowdata.append(row)

        ttk.Label(
            self.main_frame,
            text="Escribe y presiona ENTER para buscar",
            font=("Sans-serif", 10),
            justify="left",
            anchor="nw",
        ).pack(expand=True, fill="x")

        # Crear tabla
        dt = Tableview(
            master=self.main_frame,
            coldata=coldata,
            rowdata=rowdata,
            paginated=True,
            searchable=True,
            bootstyle=PRIMARY,
            autofit=True,
            autoalign=True,
        )
        dt.pack(fill=BOTH, expand=YES, padx=10, pady=10)

        # Centrar cabeceras y filas de la tabla
        for col_id in dt.view["columns"]:
            dt.view.column(col_id, anchor="center")
            dt.view.heading(col_id, anchor="center")

        # Crear el botón de exportar
        export_button = ttk.Button(
            self.main_frame,
            text="Exportar a Excel",
            command=lambda: self.export_to_excel_2(dt, fecha, grado, seccion),
            bootstyle=SUCCESS,
        )
        export_button.pack(pady=10)

    def set_alumno_add_view(self):
        """Vista para agregar un alumno"""
        self.reset_view("Crear alumno", is_expand=False, padding=15)

        form_frame = ScrolledFrame(self.wind, autohide=True)
        form_frame.pack(expand=True, fill="both")

        # codigo
        codigo_frame = ttk.Frame(form_frame)
        codigo_frame.pack(expand=True, fill="none", padx=40, pady=20)

        ttk.Label(
            codigo_frame,
            text="Código del alumno:",
            font=("Sans-Serif", 12),
        ).pack(fill="x", pady=(0, 15))

        codigo = ttk.Entry(codigo_frame, font=("Sans-Serif", 15), width=80)
        codigo.pack(fill="x")

        # nombres
        nombres_frame = ttk.Frame(form_frame)
        nombres_frame.pack(expand=True, fill="none", padx=40, pady=20)

        ttk.Label(
            nombres_frame,
            text="Nombres:",
            font=("Sans-Serif", 12),
        ).pack(fill="x", pady=(0, 15))

        nombres = ttk.Entry(nombres_frame, font=("Sans-Serif", 15), width=80)
        nombres.pack(fill="x")

        # paterno
        paterno_frame = ttk.Frame(form_frame)
        paterno_frame.pack(expand=True, fill="none", padx=40, pady=20)

        ttk.Label(
            paterno_frame,
            text="Apellido paterno:",
            font=("Sans-Serif", 12),
        ).pack(fill="x", pady=(0, 15))

        paterno = ttk.Entry(paterno_frame, font=("Sans-Serif", 15), width=80)
        paterno.pack(fill="x")

        # materno
        materno_frame = ttk.Frame(form_frame)
        materno_frame.pack(expand=True, fill="none", padx=40, pady=20)

        ttk.Label(
            materno_frame,
            text="Apellido materno:",
            font=("Sans-Serif", 12),
        ).pack(fill="x", pady=(0, 15))

        materno = ttk.Entry(materno_frame, font=("Sans-Serif", 15), width=80)
        materno.pack(fill="x")

        comboboxes_frame = ttk.Frame(form_frame)
        comboboxes_frame.pack(expand=True, padx=40, pady=20, fill="none")
        # ingreso
        ingreso_frame = ttk.Frame(comboboxes_frame)
        ingreso_frame.pack(expand=True, padx=40, pady=0, side="left")

        ttk.Label(
            ingreso_frame,
            text="Fecha de ingreso:",
            font=("Sans-Serif", 12),
        ).pack(pady=(0, 15))

        ingreso = ttk.DateEntry(
            ingreso_frame, bootstyle="primary", dateformat="%d-%m-%Y"
        )
        ingreso.pack(side="left")

        # grados
        grados_frame = ttk.Frame(comboboxes_frame)
        grados_frame.pack(expand=True, padx=40, pady=0, side="left")

        ttk.Label(
            grados_frame,
            text="Grado:",
            font=("Sans-Serif", 12),
        ).pack(pady=(0, 15))

        grados = self.run_query(
            "SELECT grado_id, grado FROM grados WHERE grado_id >= 7"
        ).fetchall()
        grados_dict = {grado: grado_id for grado_id, grado in grados}
        combobox_grados = ttk.Combobox(
            grados_frame,
            bootstyle="primary",
            values=list(grados_dict.keys()),
            state="readonly",
        )
        combobox_grados.set("Grado")
        combobox_grados.pack(side="left")

        # Secciones
        secciones_frame = ttk.Frame(comboboxes_frame)
        secciones_frame.pack(expand=True, padx=40, pady=0, side="left")

        ttk.Label(
            secciones_frame,
            text="Sección:",
            font=("Sans-Serif", 12),
        ).pack(pady=(0, 15))

        secciones = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
        combobox_secciones = ttk.Combobox(
            secciones_frame, bootstyle="primary", values=secciones, state="readonly"
        )
        combobox_secciones.set("Sección")
        combobox_secciones.pack(pady=0, padx=0, side="left")

        def set_validate_report():
            """Validar los datos y mostrar el reporte"""
            codigo_val = codigo.get()
            nombres_val = nombres.get()
            paterno_val = paterno.get()
            materno_val = materno.get()
            grado = combobox_grados.get()
            seccion = combobox_secciones.get()
            fecha = ingreso.entry.get()
            if (
                codigo_val in (None, "")
                or nombres_val in (None, "")
                or paterno_val in (None, "")
                or materno_val in (None, "")
                or grado == "Grado"
                or seccion == "Sección"
            ):
                self.display_error_box("Faltan datos en el formulario.")
                return

            self.create_alumno(
                codigo_val,
                nombres_val,
                paterno_val,
                materno_val,
                fecha,
                grados_dict[grado],
                seccion,
            )

        ttk.Button(
            form_frame,
            text="Guardar",
            bootstyle=PRIMARY,
            padding=(20, 10),
            command=set_validate_report,
        ).pack(expand=True, fill="none", pady=20)

    def set_alumnos_view(self):
        """Ver todos los alumnos"""
        self.reset_view("Alumnos", is_expand=False, padding=15)
        # Asistencias del alumno seleccionado
        alumnos = self.run_query(
            """
            SELECT
                al.alumno_id,
                al.codigo,
                al.nombres,
                al.apellido_paterno,
                al.apellido_materno,
                al.fecha_ingreso,
                g.grado,
                dg.seccion
            FROM
                alumnos al
            INNER JOIN detalle_grados dg ON
                al.detalle_grado_id = dg.detalle_grado_id
            INNER JOIN grados g ON
                dg.grado_id = g.grado_id
            ORDER BY al.alumno_id ASC
            """,
            [],
        ).fetchall()

        coldata = [
            {"text": "ID", "stretch": True},
            {"text": "Código", "stretch": True},
            {"text": "Nombres", "stretch": True},
            {"text": "Apellido paterno", "stretch": True},
            {"text": "Apellido materno", "stretch": True},
            {"text": "Fecha de ingreso", "stretch": True},
            {"text": "Grado", "stretch": True},
            {"text": "Sección", "stretch": True},
        ]

        table_frame = ttk.Frame(self.wind)
        table_frame.pack(expand=True, fill="both", padx=5, pady=5)
        table_frame.configure(width=100, height=400)

        scrolled_frame = ScrolledFrame(table_frame, autohide=False)
        scrolled_frame.pack(expand=True, fill="both")

        ttk.Label(
            scrolled_frame,
            text="- Utiliza el scroll mouse (rueda) para navegar hacia abajo en la tabla.",
            font=("Sans-serif", 10),
            justify="left",
            anchor="nw",
        ).pack(fill="x", padx=60, pady=10)

        ttk.Label(
            scrolled_frame,
            text="- Escribe y presiona ENTER para buscar.",
            font=("Sans-serif", 10),
            justify="left",
            anchor="nw",
        ).pack(fill="x", padx=60, pady=10)

        def delete_alumno():
            """Eliminar alumno"""
            selection = dt.view.selection()[0]
            row_data = dt.get_row(iid=selection).values
            alumno_id = row_data[0]

            deleted = self.run_query(
                "DELETE FROM alumnos WHERE alumno_id = ?", [alumno_id]
            )

            if deleted:
                self.display_success_toast("Alumno eliminado")
                # self.set_alumnos_view()
                return dt.delete_row(iid=selection)

            return self.display_error_box("Error interno al eliminar un alumno")

        def edit_alumno():
            """Eliminar alumno"""
            selection = dt.view.selection()[0]
            row_data = dt.get_row(iid=selection).values
            alumno_id = row_data[0]

            self.set_alumno_edit_view(alumno_id)

        buttons_frame = ttk.Frame(scrolled_frame)
        buttons_frame.pack(expand=True, fill="x", padx=60)

        ttk.Button(
            buttons_frame,
            text="Eliminar seleccionado",
            command=delete_alumno,
            bootstyle=DANGER,
        ).pack(pady=20, padx=10, fill="none", side="right")

        ttk.Button(
            buttons_frame,
            text="Editar seleccionado",
            command=edit_alumno,
            bootstyle=PRIMARY,
        ).pack(pady=20, padx=10, fill="none", side="right")

        # Crear tabla
        dt = Tableview(
            master=scrolled_frame,
            coldata=coldata,
            rowdata=alumnos,
            paginated=True,
            searchable=True,
            bootstyle=PRIMARY,
            autofit=True,
            autoalign=True,
            height=30,
            pagesize=30,
        )
        dt.pack(fill="both", expand=True, padx=60, pady=0)
        dt.config(width=80)

        # Centrar cabeceras y filas de la tabla
        for col_id in dt.view["columns"]:
            dt.view.column(col_id, anchor="center")
            dt.view.heading(col_id, anchor="center")

    def set_alumno_edit_view(self, alumno_id):
        """Editar alumno"""
        alumno = self.run_query(
            """
            SELECT
                al.alumno_id,
                al.codigo,
                al.nombres,
                al.apellido_paterno,
                al.apellido_materno,
                al.fecha_ingreso,
                g.grado,
                dg.seccion
            FROM
                alumnos al
            INNER JOIN detalle_grados dg ON
                dg.detalle_grado_id = al.detalle_grado_id
            INNER JOIN grados g ON
                dg.grado_id = g.grado_id
            WHERE
                alumno_id = ?
            """,
            [alumno_id],
        ).fetchone()
        self.reset_view("Editar alumno", is_expand=False, padding=15)
        self.set_change_view_link_corner("Regresar", self.set_alumnos_view)

        form_frame = ScrolledFrame(self.wind, autohide=True)
        form_frame.pack(expand=True, fill="both")

        # codigo
        codigo_frame = ttk.Frame(form_frame)
        codigo_frame.pack(expand=True, fill="none", padx=40, pady=20)

        ttk.Label(
            codigo_frame,
            text="Código del alumno:",
            font=("Sans-Serif", 12),
        ).pack(fill="x", pady=(0, 15))

        codigo = ttk.Entry(codigo_frame, font=("Sans-Serif", 15), width=80)
        codigo.pack(fill="x")

        # nombres
        nombres_frame = ttk.Frame(form_frame)
        nombres_frame.pack(expand=True, fill="none", padx=40, pady=20)

        ttk.Label(
            nombres_frame,
            text="Nombres:",
            font=("Sans-Serif", 12),
        ).pack(fill="x", pady=(0, 15))

        nombres = ttk.Entry(nombres_frame, font=("Sans-Serif", 15), width=80)
        nombres.pack(fill="x")

        # paterno
        paterno_frame = ttk.Frame(form_frame)
        paterno_frame.pack(expand=True, fill="none", padx=40, pady=20)

        ttk.Label(
            paterno_frame,
            text="Apellido paterno:",
            font=("Sans-Serif", 12),
        ).pack(fill="x", pady=(0, 15))

        paterno = ttk.Entry(paterno_frame, font=("Sans-Serif", 15), width=80)
        paterno.pack(fill="x")

        # materno
        materno_frame = ttk.Frame(form_frame)
        materno_frame.pack(expand=True, fill="none", padx=40, pady=20)

        ttk.Label(
            materno_frame,
            text="Apellido materno:",
            font=("Sans-Serif", 12),
        ).pack(fill="x", pady=(0, 15))

        materno = ttk.Entry(materno_frame, font=("Sans-Serif", 15), width=80)
        materno.pack(fill="x")

        comboboxes_frame = ttk.Frame(form_frame)
        comboboxes_frame.pack(expand=True, padx=40, pady=20, fill="none")
        # ingreso
        ingreso_frame = ttk.Frame(comboboxes_frame)
        ingreso_frame.pack(expand=True, padx=40, pady=0, side="left")

        ttk.Label(
            ingreso_frame,
            text="Fecha de ingreso:",
            font=("Sans-Serif", 12),
        ).pack(pady=(0, 15))

        if alumno[5] is not None:
            fecha_formatted = datetime.strptime(alumno[5], "%Y-%m-%d").date()
        else:
            fecha_formatted = datetime.now()

        ingreso = ttk.DateEntry(
            ingreso_frame,
            bootstyle="primary",
            dateformat="%d-%m-%Y",
            startdate=fecha_formatted,
        )
        ingreso.pack(side="left")

        # grados
        grados_frame = ttk.Frame(comboboxes_frame)
        grados_frame.pack(expand=True, padx=40, pady=0, side="left")

        ttk.Label(
            grados_frame,
            text="Grado:",
            font=("Sans-Serif", 12),
        ).pack(pady=(0, 15))

        grados = self.run_query(
            "SELECT grado_id, grado FROM grados WHERE grado_id >= 7"
        ).fetchall()
        grados_dict = {grado: grado_id for grado_id, grado in grados}
        combobox_grados = ttk.Combobox(
            grados_frame,
            bootstyle="primary",
            values=list(grados_dict.keys()),
            state="readonly",
        )
        combobox_grados.pack(side="left")

        # Secciones
        secciones_frame = ttk.Frame(comboboxes_frame)
        secciones_frame.pack(expand=True, padx=40, pady=0, side="left")

        ttk.Label(
            secciones_frame,
            text="Sección:",
            font=("Sans-Serif", 12),
        ).pack(pady=(0, 15))

        secciones = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
        combobox_secciones = ttk.Combobox(
            secciones_frame, bootstyle="primary", values=secciones, state="readonly"
        )
        combobox_secciones.set("Sección")
        combobox_secciones.pack(pady=0, padx=0, side="left")

        def set_validate_report():
            """Validar los datos y mostrar el reporte"""
            codigo_val = codigo.get()
            nombres_val = nombres.get()
            paterno_val = paterno.get()
            materno_val = materno.get()
            grado = combobox_grados.get()
            seccion = combobox_secciones.get()
            fecha = ingreso.entry.get()
            if (
                codigo_val in (None, "")
                or nombres_val in (None, "")
                or paterno_val in (None, "")
                or materno_val in (None, "")
                or grado == "Grado"
                or seccion == "Sección"
            ):
                self.display_error_box("Faltan datos en el formulario.")
                return

            self.update_alumno(
                alumno[0],
                codigo_val,
                nombres_val,
                paterno_val,
                materno_val,
                fecha,
                grados_dict[grado],
                seccion,
            )

        codigo.insert(0, alumno[1])
        nombres.insert(0, alumno[2])
        paterno.insert(0, alumno[3])
        materno.insert(0, alumno[4])
        combobox_grados.set(alumno[6])
        combobox_secciones.set(alumno[7])

        ttk.Button(
            form_frame,
            text="Guardar",
            bootstyle=PRIMARY,
            padding=(20, 10),
            command=set_validate_report,
        ).pack(expand=True, fill="none", pady=20)


if __name__ == "__main__":
    # Ventana principal
    window = Tk()

    # Activar modo oscuro
    # style = ttk.Style("darkly")

    app = Main(window, "escuela.db")
    window.mainloop()
